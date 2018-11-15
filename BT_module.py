from selenium import webdriver
from selenium.webdriver.support.ui import Select
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import openpyxl

def connect_site(cname):
    ide = webdriver.Firefox()
    ide.get(cname)
    ide.implicitly_wait(30)

    pop=ide.find_element_by_id("cookieNotification")
    if pop.is_displayed():
        ide.find_element_by_xpath("/html/body/div[7]/div/a[2]").click()

    ele = ide.find_element_by_id("login-username")
    ele.send_keys("GNatarajan")
    ele1 = ide.find_element_by_id("login-password")
    ele1.send_keys("1qazXSW@")

    ide.implicitly_wait(10)
    ide.find_element_by_class_name("icon-chevron-copy3").click()

    element_present = EC.visibility_of(ide.find_element_by_class_name("My-apps"))
    WebDriverWait(ide, 30).until(element_present)
    ide.find_element_by_class_name("My-apps").click()

    element_present = EC.visibility_of(ide.find_element_by_css_selector("a[href*='esales_enu_btwcom']"))
    WebDriverWait(ide, 30).until(element_present)
    ide.find_element_by_css_selector("a[href*='esales_enu_btwcom']").click()

    element_present = EC.frame_to_be_available_and_switch_to_it("_sweclient")
    WebDriverWait(ide, 30).until(element_present)
    ide._switch_to.default_content()

    wb = openpyxl.load_workbook("C:\Users\eubcefm\Desktop\BT Reference.xlsx")
    ws = wb.active
    img_list = []
    row_count=0

    for k in range(2,2000):
        if (ws.cell(row=k, column=3).value)!=None:
            row_count+=1

    for i in range(2, row_count+2):
        BT_Ref = ws.cell(row=i, column=3).value
        element_present = EC.frame_to_be_available_and_switch_to_it("_sweclient")
        WebDriverWait(ide, 30).until(element_present)
        ide.switch_to.frame("_sweview")
        select1 = ide.find_element_by_xpath("/html/body/table/tbody/tr/td/form[1]/table/tbody/tr/td[2]/table/tbody/tr[2]/td[1]/table/tbody/tr[2]/td[2]/nobr")
        element_present = EC.visibility_of(select1.find_element_by_name("s_2_1_0_0"))
        WebDriverWait(ide, 30).until(element_present)
        ide._switch_to.default_content()
        ide.switch_to.frame("_sweclient")
        ide.switch_to.frame("_sweview")
        select1 = ide.find_element_by_xpath("/html/body/table/tbody/tr/td/form[1]/table/tbody/tr/td[2]/table/tbody/tr[2]/td[1]/table/tbody/tr[2]/td[2]/nobr")
        select_box = Select(select1.find_element_by_name("s_2_1_0_0"))
        select_box.select_by_visible_text("Track a fault")
        textb = ide.find_element_by_xpath("/html/body/table/tbody/tr/td/form[1]/table/tbody/tr/td[2]/table/tbody/tr[2]/td[1]/table/tbody/tr[2]/td[3]/table/tbody/tr/td[1]/nobr")
        textb.find_element_by_id("rf").send_keys(BT_Ref.strip())
        ide.find_element_by_css_selector("a[href*='JavaScript:SWESubmitForm']").click()
        ide._switch_to.default_content()
        element_present = EC.frame_to_be_available_and_switch_to_it("_sweclient")
        WebDriverWait(ide, 30).until(element_present)
        ide.switch_to.frame("_sweview")

        img_lis1=[]
        img_lis1=ide.find_elements_by_xpath(
            "//html/body/table/tbody/tr[2]/td/form[2]/table[1]/tbody/tr[2]/td[2]/table/tbody/tr[6]/td/table/tbody/tr")
        if len(img_lis1)==0:
            img_lis1 = ide.find_elements_by_xpath(
                "//html/body/table/tbody/tr[2]/td/form[3]/table[1]/tbody/tr[2]/td[2]/table/tbody/tr[6]/td/table/tbody/tr")


        print img_lis1
        print len(img_lis1)
        tds=[]
        CRM_Count=1
        for j in img_lis1:
            if (j.get_attribute("class") == "listRowEven") or (j.get_attribute("class") == "listRowOdd"):
                tds = j.find_elements_by_tag_name('td')
                Status1=tds[3].find_element_by_tag_name('span').get_attribute('innerHTML')
                Msg1 = tds[1].find_element_by_tag_name('span').get_attribute('innerHTML')
                Status=Status1.encode("ascii")
                Msg = Msg1.encode("ascii")
                #if (Status == "CRM User Attention"):
                if (Status == "CRM User Attention") and ("Message unread" in Msg):
                    ws.cell(row=i, column=4).value = "New Update Available"
                    Update = tds[4].find_element_by_tag_name('span').get_attribute('innerHTML')
                    ws.cell(row=i, column=4+CRM_Count).value = Update
                    CRM_Count = CRM_Count + 1
        if (ws.cell(row=i, column=4).value == None):
            ws.cell(row=i, column=4).value = "No New Update Available"

        wb.save("C:\Users\eubcefm\Desktop\BT Reference.xlsx")
        ide.back()
        ide._switch_to.default_content()
        element_present1 = EC.frame_to_be_available_and_switch_to_it("_sweclient")
        WebDriverWait(ide,30).until(element_present1)
        ide._switch_to.default_content()

    element_present1 = EC.frame_to_be_available_and_switch_to_it("_sweclient")
    WebDriverWait(ide,30).until(element_present1)
    ide.switch_to.frame("_sweview")
    ide.find_element_by_id("logout").click()
